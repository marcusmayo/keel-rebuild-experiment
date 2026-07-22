"""Phase 5 -- Excel export.

Builds one workbook with four sheets: Cross-Source (matched rows),
Source-Only (gap + done_gap), Unconfirmed (unclaimed portfolio items),
Semantic (judged ambiguous rows). Shared by the CLI and the web app.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from . import config
from .portfolio import load_portfolio_map
from .reconcile import load_reconcile


def build_workbook() -> Workbook:
    data = load_reconcile()
    rows = data.get("rows", [])
    summary = data.get("summary", {})
    portfolio = load_portfolio_map()

    cross = [r for r in rows if r["bucket"] in config.MATCHED_BUCKETS]
    source_only = [r for r in rows if r["bucket"] in config.UNMATCHED_BUCKETS]
    unconfirmed_ids = summary.get("unconfirmed", [])
    semantic_rows = [r for r in rows if r["bucket"] == "ambiguous"]

    wb = Workbook()

    ws = wb.active
    ws.title = "Cross-Source"
    ws.append(["source", "ref", "title", "status", "bucket", "match_target",
               "portfolio_title", "portfolio_status", "score",
               "semantic_verdict", "semantic_reason"])
    for r in cross:
        sem = r.get("semantic") or {}
        ws.append([
            r.get("source"), r.get("ref"), r.get("title"), r.get("status"),
            r.get("bucket"), r.get("match_target"), r.get("portfolio_title"),
            r.get("portfolio_status"), r.get("score"),
            sem.get("verdict"), sem.get("reason"),
        ])

    ws2 = wb.create_sheet("Source-Only")
    ws2.append(["source", "ref", "title", "status", "bucket", "score"])
    for r in source_only:
        ws2.append([r.get("source"), r.get("ref"), r.get("title"), r.get("status"),
                    r.get("bucket"), r.get("score")])

    ws3 = wb.create_sheet("Unconfirmed")
    ws3.append(["id", "title", "status"])
    for pid in unconfirmed_ids:
        it = portfolio.get(pid, {})
        ws3.append([pid, it.get("title", ""), it.get("status", "")])

    ws4 = wb.create_sheet("Semantic")
    ws4.append(["source", "title", "status", "match_target", "portfolio_title",
                "score", "verdict", "reason", "judged"])
    for r in semantic_rows:
        sem = r.get("semantic") or {}
        ws4.append([r.get("source"), r.get("title"), r.get("status"),
                    r.get("match_target"), r.get("portfolio_title"), r.get("score"),
                    sem.get("verdict"), sem.get("reason"), sem.get("judged")])

    return wb


def export_to(path: Path | None = None) -> Path:
    if path is None:
        path = config.EXPORT_FILE
    config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(path)
    return path


if __name__ == "__main__":
    out = export_to()
    print(f"Exported workbook to {out}")
