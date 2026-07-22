"""Phase 2 -- Read-only normalization of the backlog XLSX export.

Maps each backlog row into a canonical proposal record with an empty ref
(no refs in the spreadsheet) written to proposals/backlog/<seq>.json.
Never touches state/.
"""
from __future__ import annotations

import json

from openpyxl import load_workbook

from . import config
from .generate import BACKLOG_XLSX

OUT_DIR = config.PROPOSALS_DIR / "backlog"


def normalize_backlog() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(BACKLOG_XLSX, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return 0
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    idx_title = header.index("title") if "title" in header else 0
    idx_status = header.index("status") if "status" in header else 1
    count = 0
    for seq, row in enumerate(rows[1:], start=1):
        title = (row[idx_title] if idx_title < len(row) else "") or ""
        status = (row[idx_status] if idx_status < len(row) else "") or ""
        record = {
            "source": "backlog",
            "seq": seq,
            "ref": "",
            "title": str(title).strip(),
            "status": str(status).strip(),
        }
        with open(OUT_DIR / f"{seq:04d}.json", "w") as out:
            json.dump(record, out, indent=2, sort_keys=True, ensure_ascii=False)
            out.write("\n")
        count += 1
    return count


if __name__ == "__main__":
    print(f"Normalized {normalize_backlog()} backlog proposals.")
