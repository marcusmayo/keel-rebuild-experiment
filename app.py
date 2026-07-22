#!/usr/bin/env python3
"""Phase 5: Operator webchat — Flask app with TOTP login, dashboard, reconciliation review."""
import os
import io
import json
import yaml
import time
import pyotp
import secrets
from datetime import datetime
from flask import Flask, request, jsonify, send_file, session, redirect, url_for, render_template_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

STATE_DIR = "state"
LOGS_DIR = "logs"
SECRET_FILE = ".totp_secret"

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(32))


# ── TOTP ────────────────────────────────────────────────────────────────────
def load_totp_secret():
    if not os.path.exists(SECRET_FILE):
        # Auto-generate if missing
        secret = pyotp.random_base32()
        with open(SECRET_FILE, "w") as f:
            f.write(secret)
    with open(SECRET_FILE) as f:
        return f.read().strip()


TOTP_SECRET = load_totp_secret()
totp = pyotp.TOTP(TOTP_SECRET)


# ── Data loaders ────────────────────────────────────────────────────────────
def load_portfolio():
    items = {}
    for fname in sorted(os.listdir(STATE_DIR)):
        if fname.endswith((".yaml", ".yml")):
            with open(os.path.join(STATE_DIR, fname)) as f:
                item = yaml.safe_load(f)
            items[item["id"]] = item
    return items


def load_reconcile():
    if not os.path.exists("reconcile.json"):
        return {"summary": {}, "unconfirmed": [], "results": []}
    with open("reconcile.json") as f:
        return json.load(f)


# ── Auth decorator ──────────────────────────────────────────────────────────
def login_required(f):
    from functools import wraps

    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("authenticated"):
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)

    return decorated


# ── Routes ──────────────────────────────────────────────────────────────────
@app.route("/")
def login_page():
    if session.get("authenticated"):
        return redirect(url_for("dashboard"))
    return render_template_string(LOGIN_HTML)


@app.route("/login", methods=["POST"])
def login():
    data = request.get_json() or {}
    code = data.get("code", "")
    if totp.verify(code):
        session["authenticated"] = True
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "Invalid TOTP code"}), 401


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))


@app.route("/dashboard")
@login_required
def dashboard():
    reconcile = load_reconcile()
    summary = reconcile.get("summary", {})
    unconfirmed = reconcile.get("unconfirmed", [])
    return render_template_string(
        DASHBOARD_HTML,
        summary=summary,
        unconfirmed_count=len(unconfirmed),
        unconfirmed_list=", ".join(unconfirmed),
    )


@app.route("/api/dashboard")
@login_required
def api_dashboard():
    reconcile = load_reconcile()
    return jsonify({
        "summary": reconcile.get("summary", {}),
        "unconfirmed": reconcile.get("unconfirmed", []),
        "unconfirmed_count": len(reconcile.get("unconfirmed", [])),
    })


@app.route("/items")
@login_required
def items_page():
    portfolio = load_portfolio()
    reconcile = load_reconcile()
    results = reconcile.get("results", [])

    # Map which proposals match each portfolio item
    match_map = {}
    for r in results:
        if r.get("matched_id") and r.get("bucket") != "duplicate":
            pid = r["matched_id"]
            if pid not in match_map:
                match_map[pid] = []
            match_map[pid].append(r)

    items_data = []
    for pid, pitem in sorted(portfolio.items()):
        matches = match_map.get(pid, [])
        items_data.append({
            "id": pid,
            "title": pitem["title"],
            "status": pitem["status"],
            "source_matches": len(matches),
            "buckets": [m["bucket"] for m in matches],
        })
    return render_template_string(ITEMS_HTML, items=items_data)


@app.route("/api/items")
@login_required
def api_items():
    portfolio = load_portfolio()
    reconcile = load_reconcile()
    results = reconcile.get("results", [])

    match_map = {}
    for r in results:
        if r.get("matched_id") and r.get("bucket") != "duplicate":
            pid = r["matched_id"]
            if pid not in match_map:
                match_map[pid] = []
            match_map[pid].append(r)

    items_data = []
    for pid, pitem in sorted(portfolio.items()):
        matches = match_map.get(pid, [])
        items_data.append({
            "id": pid,
            "title": pitem["title"],
            "status": pitem["status"],
            "source_matches": len(matches),
            "buckets": [m["bucket"] for m in matches],
        })
    return jsonify(items_data)


@app.route("/review")
@login_required
def review_page():
    reconcile = load_reconcile()
    results = reconcile.get("results", [])
    summary = reconcile.get("summary", {})

    # Group by bucket
    buckets = {}
    for r in results:
        b = r["bucket"]
        if b not in buckets:
            buckets[b] = []
        buckets[b].append(r)

    return render_template_string(
        REVIEW_HTML,
        buckets=buckets,
        summary=summary,
        bucket_order=["changed", "conflict", "completed", "duplicate", "ambiguous", "gap", "done_gap"],
    )


@app.route("/api/review")
@login_required
def api_review():
    reconcile = load_reconcile()
    return jsonify(reconcile)


@app.route("/api/items/<item_id>")
@login_required
def api_item_detail(item_id):
    portfolio = load_portfolio()
    if item_id not in portfolio:
        return jsonify({"error": "not found"}), 404
    pitem = portfolio[item_id]
    reconcile = load_reconcile()
    matching_rows = [
        r for r in reconcile.get("results", [])
        if r.get("matched_id") == item_id and r.get("bucket") != "duplicate"
    ]
    return jsonify({
        "portfolio": pitem,
        "matches": matching_rows,
    })


@app.route("/api/items/<item_id>/propose_status", methods=["POST"])
@login_required
def propose_status(item_id):
    """Propose a status change. Returns a confirmation token."""
    data = request.get_json() or {}
    new_status = data.get("status", "").strip()
    if not new_status:
        return jsonify({"error": "status required"}), 400

    portfolio = load_portfolio()
    if item_id not in portfolio:
        return jsonify({"error": "not found"}), 404

    pitem = portfolio[item_id]
    old_status = pitem["status"]

    if old_status == new_status:
        return jsonify({"error": "no change"}), 400

    # Store proposal in session
    token = secrets.token_hex(8)
    session[f"proposal_{token}"] = {
        "item_id": item_id,
        "old_status": old_status,
        "new_status": new_status,
        "timestamp": time.time(),
    }

    return jsonify({
        "ok": True,
        "token": token,
        "item_id": item_id,
        "old_status": old_status,
        "new_status": new_status,
        "message": f"Proposed: {item_id} status {old_status} → {new_status}. Confirm to apply.",
    })


@app.route("/api/items/<item_id>/confirm_status", methods=["POST"])
@login_required
def confirm_status(item_id):
    """Confirm a proposed status change."""
    data = request.get_json() or {}
    token = data.get("token", "").strip()
    if not token:
        return jsonify({"error": "confirmation token required"}), 400

    proposal = session.pop(f"proposal_{token}", None)
    if not proposal:
        return jsonify({"error": "invalid or expired token"}), 400

    if proposal["item_id"] != item_id:
        return jsonify({"error": "token mismatch"}), 400

    # Apply the change
    path = os.path.join(STATE_DIR, f"{item_id}.yaml")
    with open(path) as f:
        item = yaml.safe_load(f)
    item["status"] = proposal["new_status"]
    with open(path, "w") as f:
        yaml.dump(item, f, default_flow_style=False, sort_keys=False)

    # Audit log
    os.makedirs(LOGS_DIR, exist_ok=True)
    audit_line = (
        f"{datetime.utcnow().isoformat()}Z | {item_id} | "
        f"{proposal['old_status']} → {proposal['new_status']} | "
        f"confirmed\n"
    )
    with open(os.path.join(LOGS_DIR, "audit.log"), "a") as f:
        f.write(audit_line)

    return jsonify({
        "ok": True,
        "item_id": item_id,
        "old_status": proposal["old_status"],
        "new_status": proposal["new_status"],
        "audit_logged": True,
    })


@app.route("/export")
@login_required
def export_xlsx():
    """Download Excel workbook with four sheets."""
    reconcile = load_reconcile()
    results = reconcile.get("results", [])
    portfolio = load_portfolio()

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B3A4E", end_color="2B3A4E", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def style_data(ws, rows, start_row=2):
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border

    # Sheet 1: Cross-Source (matched rows: changed, conflict, completed, duplicate)
    ws1 = wb.create_sheet("Cross-Source")
    headers1 = ["Source", "Ref", "Title", "Status", "Bucket", "Matched ID", "Score"]
    style_header(ws1, headers1)
    matched_rows = [
        [r["source"], r["ref"], r["title"], r["status"], r["bucket"],
         r.get("matched_id", ""), r.get("match_score", "")]
        for r in results
        if r["bucket"] in ("changed", "conflict", "completed", "duplicate")
    ]
    style_data(ws1, matched_rows)
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = 22

    # Sheet 2: Source-Only (gap + done_gap)
    ws2 = wb.create_sheet("Source-Only")
    headers2 = ["Source", "Ref", "Title", "Status", "Bucket", "Best Match", "Score"]
    style_header(ws2, headers2)
    source_only = [
        [r["source"], r["ref"], r["title"], r["status"], r["bucket"],
         r.get("matched_id", ""), r.get("match_score", "")]
        for r in results
        if r["bucket"] in ("gap", "done_gap")
    ]
    style_data(ws2, source_only)
    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = 22

    # Sheet 3: Unconfirmed (portfolio items no source claims)
    ws3 = wb.create_sheet("Unconfirmed")
    headers3 = ["ID", "Title", "Status"]
    style_header(ws3, headers3)
    unconfirmed = reconcile.get("unconfirmed", [])
    unconfirmed_rows = [
        [pid, portfolio[pid]["title"], portfolio[pid]["status"]]
        for pid in unconfirmed
        if pid in portfolio
    ]
    style_data(ws3, unconfirmed_rows)
    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[ws3.cell(row=1, column=col).column_letter].width = 30

    # Sheet 4: Semantic (ambiguous/judged rows)
    ws4 = wb.create_sheet("Semantic")
    headers4 = ["Source", "Title", "Bucket", "Best Match", "Score", "Verdict", "Reason"]
    style_header(ws4, headers4)
    semantic_rows = [
        [r["source"], r["title"], r["bucket"], r.get("matched_id", ""),
         r.get("match_score", ""), r.get("semantic_verdict", ""),
         r.get("semantic_reason", "")]
        for r in results
        if r["bucket"] == "ambiguous"
    ]
    style_data(ws4, semantic_rows)
    for col in range(1, len(headers4) + 1):
        ws4.column_dimensions[ws4.cell(row=1, column=col).column_letter].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="meridian_reconciliation.xlsx",
    )


# ── HTML Templates ──────────────────────────────────────────────────────────

LOGIN_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Meridian PPM — Login</title>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: 'SF Mono', 'Fira Code', 'Consolas', monospace; background: #0D1117; color: #C9D1D9; display: flex; align-items: center; justify-content: center; min-height: 100vh; }
  .login-box { background: #161B22; border: 1px solid #30363D; border-radius: 8px; padding: 40px; width: 400px; max-width: 90vw; }
  h1 { color: #58A6FF; font-size: 20px; margin-bottom: 8px; }
  .subtitle { color: #8B949E; font-size: 13px; margin-bottom: 24px; }
  input { width: 100%; padding: 10px 12px; background: #0D1117; border: 1px solid #30363D; border-radius: 6px; color: #C9D1D9; font-family: inherit; font-size: 16px; text-align: center; letter-spacing: 4px; }
  input:focus { outline: none; border-color: #58A6FF; }
  button { width: 100%; margin-top: 16px; padding: 10px; background: #238636; border: none; border-radius: 6px; color: white; font-family: inherit; font-size: 14px; cursor: pointer; }
  button:hover { background: #2EA043; }
  .error { color: #F85149; font-size: 13px; margin-top: 12px; text-align: center; display: none; }
</style>
</head>
<body>
<div class="login-box">
  <h1>Meridian PPM</h1>
  <p class="subtitle">Portfolio reconciliation cockpit. Enter your TOTP code.</p>
  <input type="text" id="code" placeholder="000000" maxlength="6" autofocus>
  <button onclick="doLogin()">Authenticate</button>
  <p class="error" id="error">Invalid code</p>
</div>
<script>
async function doLogin() {
  const code = document.getElementById('code').value;
  const resp = await fetch('/login', { method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify({code}) });
  if (resp.ok) { window.location = '/dashboard'; }
  else { document.getElementById('error').style.display = 'block'; }
}
document.getElementById('code').addEventListener('keydown', e => { if (e.key === 'Enter') doLogin(); });
</script>
</body>
</html>"""

DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Meridian PPM — Dashboard</title>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: 'SF Mono', 'Fira Code', 'Consolas', monospace; background: #0D1117; color: #C9D1D9; }
  .nav { background: #161B22; border-bottom: 1px solid #30363D; padding: 10px 24px; display: flex; align-items: center; justify-content: space-between; }
  .nav h1 { color: #58A6FF; font-size: 16px; }
  .nav a { color: #8B949E; text-decoration: none; margin-left: 20px; font-size: 13px; }
  .nav a:hover { color: #C9D1D9; }
  .nav a.active { color: #58A6FF; }
  .container { max-width: 1200px; margin: 0 auto; padding: 24px; }
  .hero-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 12px; margin-bottom: 32px; }
  .hero-card { background: #161B22; border: 1px solid #30363D; border-radius: 8px; padding: 16px; text-align: center; }
  .hero-card .count { font-size: 36px; font-weight: bold; }
  .hero-card .label { font-size: 12px; color: #8B949E; margin-top: 4px; text-transform: uppercase; letter-spacing: 1px; }
  .changed .count { color: #D29922; } .gap .count { color: #F85149; } .conflict .count { color: #F85149; }
  .ambiguous .count { color: #A371F7; } .duplicate .count { color: #8B949E; } .completed .count { color: #3FB950; }
  .done-gap .count { color: #8B949E; } .unconfirmed .count { color: #79C0FF; }
</style>
</head>
<body>
<div class="nav">
  <h1>Meridian PPM</h1>
  <div>
    <a href="/dashboard" class="active">Dashboard</a>
    <a href="/items">Work Items</a>
    <a href="/review">Reconciliation</a>
    <a href="/export">Export</a>
    <a href="/logout">Logout</a>
  </div>
</div>
<div class="container">
  <div class="hero-grid">
    {% set buckets = [('changed', 'Changed', '#D29922'), ('gap', 'Gap', '#F85149'), ('conflict', 'Conflict', '#F85149'), ('ambiguous', 'Ambiguous', '#A371F7'), ('duplicate', 'Duplicate', '#8B949E'), ('completed', 'Completed', '#3FB950'), ('done_gap', 'Done Gap', '#8B949E')] %}
    {% for key, label, color in buckets %}
    <div class="hero-card {{ key|replace('_','-') }}">
      <div class="count" style="color:{{ color }}">{{ summary.get(key, 0) }}</div>
      <div class="label">{{ label }}</div>
    </div>
    {% endfor %}
    <div class="hero-card unconfirmed">
      <div class="count">{{ unconfirmed_count }}</div>
      <div class="label">Unconfirmed</div>
    </div>
  </div>
  <p style="color:#8B949E;font-size:12px;">Unconfirmed items: {{ unconfirmed_list }}</p>
</div>
</body>
</html>"""

ITEMS_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Meridian PPM — Work Items</title>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: 'SF Mono', 'Fira Code', 'Consolas', monospace; background: #0D1117; color: #C9D1D9; }
  .nav { background: #161B22; border-bottom: 1px solid #30363D; padding: 10px 24px; display: flex; align-items: center; justify-content: space-between; }
  .nav h1 { color: #58A6FF; font-size: 16px; }
  .nav a { color: #8B949E; text-decoration: none; margin-left: 20px; font-size: 13px; }
  .nav a:hover { color: #C9D1D9; }
  .nav a.active { color: #58A6FF; }
  .container { max-width: 1200px; margin: 0 auto; padding: 24px; }
  table { width: 100%; border-collapse: collapse; }
  th { text-align: left; padding: 10px 12px; border-bottom: 1px solid #30363D; color: #8B949E; font-size: 12px; text-transform: uppercase; letter-spacing: 1px; }
  td { padding: 10px 12px; border-bottom: 1px solid #21262D; font-size: 13px; }
  tr:hover { background: #161B22; }
  .status { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; }
  .status.in_progress { background: #1A3A5C; color: #58A6FF; }
  .status.done { background: #1A3A2A; color: #3FB950; }
  .status.blocked { background: #3A1A1A; color: #F85149; }
  .item-row { cursor: pointer; }
  .modal-overlay { display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.7); z-index: 100; align-items: center; justify-content: center; }
  .modal-overlay.show { display: flex; }
  .modal { background: #161B22; border: 1px solid #30363D; border-radius: 8px; padding: 24px; width: 500px; max-width: 90vw; }
  .modal h2 { color: #58A6FF; margin-bottom: 16px; }
  .modal select, .modal button { padding: 8px 12px; margin: 4px; border-radius: 6px; font-family: inherit; }
  .modal select { background: #0D1117; border: 1px solid #30363D; color: #C9D1D9; }
  .modal button.primary { background: #238636; border: none; color: white; cursor: pointer; }
  .modal button.secondary { background: #30363D; border: none; color: #C9D1D9; cursor: pointer; }
  .modal .msg { margin-top: 12px; font-size: 13px; }
  .modal .msg.success { color: #3FB950; }
  .modal .msg.error { color: #F85149; }
</style>
</head>
<body>
<div class="nav">
  <h1>Meridian PPM</h1>
  <div>
    <a href="/dashboard">Dashboard</a>
    <a href="/items" class="active">Work Items</a>
    <a href="/review">Reconciliation</a>
    <a href="/export">Export</a>
    <a href="/logout">Logout</a>
  </div>
</div>
<div class="container">
  <table>
    <tr><th>ID</th><th>Title</th><th>Status</th><th>Source Matches</th><th>Buckets</th></tr>
    {% for item in items %}
    <tr class="item-row" onclick="openItem('{{ item.id }}')">
      <td><code>{{ item.id }}</code></td>
      <td>{{ item.title }}</td>
      <td><span class="status {{ item.status|replace('_', '-') }}">{{ item.status }}</span></td>
      <td>{{ item.source_matches }}</td>
      <td>{{ item.buckets|join(', ') }}</td>
    </tr>
    {% endfor %}
  </table>
</div>

<div class="modal-overlay" id="modal">
  <div class="modal">
    <h2 id="modal-title">Item Detail</h2>
    <div id="modal-body"></div>
    <div id="modal-msg" class="msg"></div>
    <br>
    <button class="secondary" onclick="closeModal()">Close</button>
  </div>
</div>

<script>
let confirmToken = null;

async function openItem(id) {
  const resp = await fetch('/api/items/' + id);
  const data = await resp.json();
  document.getElementById('modal-title').textContent = data.portfolio.id + ': ' + data.portfolio.title;
  let html = '<p>Current Status: <strong>' + data.portfolio.status + '</strong></p>';
  html += '<p>Matching source rows: ' + data.matches.length + '</p>';
  html += '<p>Change status to: ';
  html += '<select id="new-status">';
  for (const s of ['in_progress', 'blocked', 'done']) {
    html += '<option value="' + s + '"' + (s === data.portfolio.status ? ' selected' : '') + '>' + s + '</option>';
  }
  html += '</select>';
  html += ' <button class="primary" onclick="proposeChange(\'' + id + '\')">Propose</button>';
  html += ' <button class="primary" id="confirm-btn" style="display:none" onclick="confirmChange(\'' + id + '\')">Confirm</button></p>';
  document.getElementById('modal-body').innerHTML = html;
  document.getElementById('modal-msg').textContent = '';
  document.getElementById('modal-msg').className = 'msg';
  document.getElementById('confirm-btn').style.display = 'none';
  confirmToken = null;
  document.getElementById('modal').classList.add('show');
}

async function proposeChange(id) {
  const newStatus = document.getElementById('new-status').value;
  const resp = await fetch('/api/items/' + id + '/propose_status', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({status: newStatus})
  });
  const data = await resp.json();
  const msgEl = document.getElementById('modal-msg');
  if (data.ok) {
    confirmToken = data.token;
    msgEl.textContent = data.message;
    msgEl.className = 'msg success';
    document.getElementById('confirm-btn').style.display = 'inline-block';
  } else {
    msgEl.textContent = data.error || 'Proposal failed';
    msgEl.className = 'msg error';
  }
}

async function confirmChange(id) {
  const resp = await fetch('/api/items/' + id + '/confirm_status', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({token: confirmToken})
  });
  const data = await resp.json();
  const msgEl = document.getElementById('modal-msg');
  if (data.ok) {
    msgEl.textContent = 'Status changed: ' + data.old_status + ' → ' + data.new_status + '. Audit logged.';
    msgEl.className = 'msg success';
    document.getElementById('confirm-btn').style.display = 'none';
    confirmToken = null;
    setTimeout(() => { closeModal(); location.reload(); }, 1500);
  } else {
    msgEl.textContent = data.error || 'Confirm failed';
    msgEl.className = 'msg error';
  }
}

function closeModal() {
  document.getElementById('modal').classList.remove('show');
}
</script>
</body>
</html>"""

REVIEW_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Meridian PPM — Reconciliation Review</title>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: 'SF Mono', 'Fira Code', 'Consolas', monospace; background: #0D1117; color: #C9D1D9; }
  .nav { background: #161B22; border-bottom: 1px solid #30363D; padding: 10px 24px; display: flex; align-items: center; justify-content: space-between; }
  .nav h1 { color: #58A6FF; font-size: 16px; }
  .nav a { color: #8B949E; text-decoration: none; margin-left: 20px; font-size: 13px; }
  .nav a:hover { color: #C9D1D9; }
  .nav a.active { color: #58A6FF; }
  .container { max-width: 1200px; margin: 0 auto; padding: 24px; }
  .bucket-section { margin-bottom: 24px; }
  .bucket-header { padding: 8px 12px; border-radius: 6px 6px 0 0; font-weight: bold; font-size: 14px; display: flex; justify-content: space-between; }
  .bucket-header.changed { background: #D2992233; color: #D29922; }
  .bucket-header.conflict { background: #F8514933; color: #F85149; }
  .bucket-header.completed { background: #3FB95033; color: #3FB950; }
  .bucket-header.duplicate { background: #8B949E33; color: #8B949E; }
  .bucket-header.ambiguous { background: #A371F733; color: #A371F7; }
  .bucket-header.gap { background: #F8514933; color: #F85149; }
  .bucket-header.done_gap { background: #8B949E33; color: #8B949E; }
  table { width: 100%; border-collapse: collapse; }
  th { text-align: left; padding: 8px 12px; border-bottom: 1px solid #30363D; color: #8B949E; font-size: 12px; text-transform: uppercase; }
  td { padding: 8px 12px; border-bottom: 1px solid #21262D; font-size: 13px; }
  tr:hover { background: #161B22; }
  .verdict-tag { display: inline-block; padding: 2px 6px; border-radius: 4px; font-size: 11px; }
  .verdict-tag.SAME { background: #3FB95033; color: #3FB950; }
  .verdict-tag.DISTINCT { background: #F8514933; color: #F85149; }
  .verdict-tag.SKIPPED { background: #8B949E33; color: #8B949E; }
</style>
</head>
<body>
<div class="nav">
  <h1>Meridian PPM</h1>
  <div>
    <a href="/dashboard">Dashboard</a>
    <a href="/items">Work Items</a>
    <a href="/review" class="active">Reconciliation</a>
    <a href="/export">Export</a>
    <a href="/logout">Logout</a>
  </div>
</div>
<div class="container">
{% for bucket_name in bucket_order %}
{% set rows = buckets.get(bucket_name, []) %}
{% if rows %}
<div class="bucket-section">
  <div class="bucket-header {{ bucket_name }}">
    <span>{{ bucket_name|replace('_', ' ')|title }}</span>
    <span>{{ rows|length }} row(s)</span>
  </div>
  <table>
    <tr><th>Source</th><th>Ref</th><th>Title</th><th>Status</th><th>Matched ID</th><th>Score</th>
    {% if bucket_name == 'ambiguous' %}<th>Verdict</th><th>Reason</th>{% endif %}
    </tr>
    {% for r in rows %}
    <tr>
      <td>{{ r.source }}</td>
      <td><code>{{ r.ref }}</code></td>
      <td>{{ r.title }}</td>
      <td>{{ r.status }}</td>
      <td><code>{{ r.matched_id or '' }}</code></td>
      <td>{{ r.match_score }}</td>
      {% if bucket_name == 'ambiguous' %}
      <td><span class="verdict-tag {{ r.get('semantic_verdict', '') }}">{{ r.get('semantic_verdict', '—') }}</span></td>
      <td style="max-width:300px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="{{ r.get('semantic_reason', '') }}">{{ r.get('semantic_reason', '—') }}</td>
      {% endif %}
    </tr>
    {% endfor %}
  </table>
</div>
{% endif %}
{% endfor %}
</div>
</body>
</html>"""

if __name__ == "__main__":
    print(f"TOTP secret: {TOTP_SECRET}")
    print(f"OTPAuth URL: {totp.provisioning_uri(name='operator', issuer_name='MeridianPPM')}")
    print(f"Current TOTP code: {totp.now()}")
    app.run(host="0.0.0.0", port=8000, debug=False)