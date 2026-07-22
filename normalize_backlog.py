"""Phase 2 -- normalize the backlog XLSX export into canonical proposals.

Read-only with respect to state/: only writes proposals/backlog.json.
Backlog rows carry no refs, so ref is the empty string.
Canonical record: {ref, title, status, source}.
"""
import json
import os

from openpyxl import load_workbook

from common import IMPORTS_DIR, PROPOSALS_DIR


def main():
    src = os.path.join(IMPORTS_DIR, "backlog_export.xlsx")
    wb = load_workbook(src, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() for c in rows[0]]
    ti, si = header.index("title"), header.index("status")
    proposals = []
    for r in rows[1:]:
        if r is None or all(c is None for c in r):
            continue
        proposals.append({
            "ref": "",
            "title": str(r[ti]).strip(),
            "status": str(r[si]).strip(),
            "source": "backlog",
        })
    wb.close()
    os.makedirs(PROPOSALS_DIR, exist_ok=True)
    out = os.path.join(PROPOSALS_DIR, "backlog.json")
    with open(out, "w") as f:
        json.dump(proposals, f, indent=2)
        f.write("\n")
    print(f"Wrote {len(proposals)} backlog proposals to {out}")


if __name__ == "__main__":
    main()
