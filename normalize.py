#!/usr/bin/env python3
"""Phase 2: Read-only normalizers — map each export into canonical proposal records."""

import os
import csv
import json
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))


def normalize_jira():
    """Read imports/jira_export.csv, emit proposals/jira_proposals.json."""
    jira_path = os.path.join(BASE, "imports", "jira_export.csv")
    proposals = []
    with open(jira_path, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            proposals.append({
                "ref": row.get("ref", "").strip(),
                "title": row.get("title", "").strip(),
                "status": row.get("status", "").strip(),
                "source": "jira",
            })
    out_dir = os.path.join(BASE, "proposals")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "jira_proposals.json")
    with open(out_path, "w") as f:
        json.dump(proposals, f, indent=2)
    print(f"Normalized {len(proposals)} Jira proposals → proposals/jira_proposals.json")
    return proposals


def normalize_backlog():
    """Read imports/backlog_export.xlsx, emit proposals/backlog_proposals.json."""
    backlog_path = os.path.join(BASE, "imports", "backlog_export.xlsx")
    wb = load_workbook(backlog_path, read_only=True)
    ws = wb.active
    proposals = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:  # header
            continue
        title = (row[0] or "").strip()
        status = (row[1] or "").strip()
        proposals.append({
            "ref": "",  # Backlog has no refs
            "title": title,
            "status": status,
            "source": "backlog",
        })
    wb.close()
    out_dir = os.path.join(BASE, "proposals")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "backlog_proposals.json")
    with open(out_path, "w") as f:
        json.dump(proposals, f, indent=2)
    print(f"Normalized {len(proposals)} Backlog proposals → proposals/backlog_proposals.json")
    return proposals


def main():
    normalize_jira()
    normalize_backlog()


if __name__ == "__main__":
    main()
