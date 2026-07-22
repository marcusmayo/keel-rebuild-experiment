#!/usr/bin/env python3
"""Phase 2 normalizers: map each export into canonical proposal records.

Two read-only normalizers turn the Jira CSV and backlog XLSX into canonical
proposal records with fields (ref, title, status, source). Output goes to
proposals/. Nothing under state/ is ever created, modified, or removed.

Canonical proposal record:
    {"ref": "ML-001" | "", "title": str, "status": str, "source": "jira"|"backlog"}
"""
from __future__ import annotations

import csv
import json
import sys
from pathlib import Path
from typing import Dict, List

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from meridian.config import (  # noqa: E402
    JIRA_CSV,
    BACKLOG_XLSX,
    JIRA_PROPOSALS,
    BACKLOG_PROPOSALS,
    ensure_dirs,
)


def normalize_jira() -> List[Dict]:
    """Read the Jira CSV into canonical proposals (source=jira)."""
    proposals: List[Dict] = []
    with JIRA_CSV.open("r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            proposals.append(
                {
                    "ref": (row.get("ref") or "").strip(),
                    "title": (row.get("title") or "").strip(),
                    "status": (row.get("status") or "").strip(),
                    "source": "jira",
                }
            )
    return proposals


def normalize_backlog() -> List[Dict]:
    """Read the backlog XLSX into canonical proposals (source=backlog).

    The backlog has no ref column, so every row carries an empty ref.
    """
    proposals: List[Dict] = []
    wb = load_workbook(BACKLOG_XLSX, read_only=True, data_only=True)
    ws = wb.active
    header: List[str] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            header = [str(c).strip().lower() if c is not None else "" for c in row]
            continue
        if row is None or all(c is None for c in row):
            continue
        record = {header[i]: row[i] for i in range(min(len(header), len(row)))}
        title = record.get("title")
        status = record.get("status")
        proposals.append(
            {
                "ref": "",
                "title": (str(title).strip() if title is not None else ""),
                "status": (str(status).strip() if status is not None else ""),
                "source": "backlog",
            }
        )
    wb.close()
    return proposals


def _write(path: Path, proposals: List[Dict]) -> None:
    with path.open("w", encoding="utf-8") as fh:
        json.dump(proposals, fh, indent=2, sort_keys=True)
        fh.write("\n")


def main() -> int:
    ensure_dirs()
    jira = normalize_jira()
    backlog = normalize_backlog()
    _write(JIRA_PROPOSALS, jira)
    _write(BACKLOG_PROPOSALS, backlog)
    print(f"Jira normalizer emitted {len(jira)} proposals -> {JIRA_PROPOSALS}")
    print(f"Backlog normalizer emitted {len(backlog)} proposals -> {BACKLOG_PROPOSALS}")
    print("normalize: OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
