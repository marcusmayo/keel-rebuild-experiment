#!/usr/bin/env python3
"""Phase 5 (standalone): Generate the Excel export from reconcile.json."""

import os
import json
import io
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

BASE = os.path.dirname(os.path.abspath(__file__))


def export_xlsx(output_path=None):
    recon_path = os.path.join(BASE, "reconcile.json")
    if not os.path.exists(recon_path):
        print("reconcile.json not found. Run reconcile first.")
        return None

    with open(recon_path) as f:
        recon = json.load(f)

    results = recon["results"]

    # Load portfolio titles for unconfirmed
    import yaml
    portfolio = {}
    state_dir = os.path.join(BASE, "state")
    for fname in os.listdir(state_dir):
        if fname.endswith(".yaml") or fname.endswith(".yml"):
            with open(os.path.join(state_dir, fname)) as f:
                item = yaml.safe_load(f)
            portfolio[item["id"]] = item

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # --- Sheet 1: Cross-Source (matched rows — not gap, not done_gap) ---
    ws1 = wb.create_sheet("Cross-Source")
    ws1.append(["Ref", "Title", "Status", "Source", "Bucket", "Match Target", "Score",
                "Semantic Verdict", "Semantic Reason"])
    for r in results:
        if r["bucket"] not in ("gap", "done_gap", "ambiguous"):
            ws1.append([
                r.get("ref", ""), r["title"], r["status"], r["source"], r["bucket"],
                r.get("match_target", ""), r.get("match_score", ""),
                r.get("semantic_verdict", ""), r.get("semantic_reason", "")
            ])
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font

    # --- Sheet 2: Source-Only (gap + done_gap) ---
    ws2 = wb.create_sheet("Source-Only")
    ws2.append(["Ref", "Title", "Status", "Source", "Bucket"])
    for r in results:
        if r["bucket"] in ("gap", "done_gap"):
            ws2.append([r.get("ref", ""), r["title"], r["status"], r["source"], r["bucket"]])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    # --- Sheet 3: Unconfirmed ---
    ws3 = wb.create_sheet("Unconfirmed")
    ws3.append(["ID", "Title", "Status"])
    unconfirmed_ids = recon["summary"].get("unconfirmed", [])
    for uid in unconfirmed_ids:
        if uid in portfolio:
            item = portfolio[uid]
            ws3.append([item["id"], item["title"], item["status"]])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font

    # --- Sheet 4: Semantic ---
    ws4 = wb.create_sheet("Semantic")
    ws4.append(["Ref", "Title", "Status", "Source", "Bucket", "Match Target", "Score",
                "Semantic Verdict", "Semantic Reason"])
    for r in results:
        if r.get("semantic_verdict"):
            ws4.append([
                r.get("ref", ""), r["title"], r["status"], r["source"], r["bucket"],
                r.get("match_target", ""), r.get("match_score", ""),
                r.get("semantic_verdict", ""), r.get("semantic_reason", "")
            ])
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font

    if output_path is None:
        output_path = os.path.join(BASE, "exports", "meridian_reconciliation.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Export written to {output_path}")
    return output_path


if __name__ == "__main__":
    export_xlsx()