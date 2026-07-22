"""Excel export: one workbook, four sheets.

Sheets (row counts for the seeded world in parentheses):
  - Cross-Source (15): the matched rows -- changed, conflict, completed,
    duplicate. Every source row whose best match is an actual portfolio item.
  - Source-Only (6):   the gap + done_gap rows (matched nothing).
  - Unconfirmed (6):   portfolio items no source row claimed.
  - Semantic (1):      the ambiguous rows carrying a model verdict.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font

from .portfolio import get_item, load_portfolio

# Buckets whose rows matched an actual portfolio item.
MATCHED_BUCKETS = ("changed", "conflict", "completed", "duplicate")
SOURCE_ONLY_BUCKETS = ("gap", "done_gap")

_HEADER_FONT = Font(bold=True)


def _write_header(ws, headers: List[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT


def build_workbook(report: Dict) -> Workbook:
    """Build the four-sheet workbook from a reconciliation report dict."""
    rows = report.get("rows", [])
    portfolio = {it["id"]: it for it in load_portfolio()}

    wb = Workbook()

    # --- Cross-Source: matched rows ---
    ws1 = wb.active
    ws1.title = "Cross-Source"
    _write_header(ws1, ["source", "ref", "title", "status", "bucket",
                        "target", "score"])
    for r in rows:
        if r["bucket"] in MATCHED_BUCKETS:
            ws1.append([r["source"], r["ref"], r["title"], r["status"],
                        r["bucket"], r["target"], r["score"]])

    # --- Source-Only: gap + done_gap ---
    ws2 = wb.create_sheet("Source-Only")
    _write_header(ws2, ["source", "ref", "title", "status", "bucket"])
    for r in rows:
        if r["bucket"] in SOURCE_ONLY_BUCKETS:
            ws2.append([r["source"], r["ref"], r["title"], r["status"],
                        r["bucket"]])

    # --- Unconfirmed: portfolio items claimed by no source row ---
    ws3 = wb.create_sheet("Unconfirmed")
    _write_header(ws3, ["id", "title", "status"])
    for item_id in report.get("summary", {}).get("unconfirmed", []):
        item = portfolio.get(item_id) or get_item(item_id) or {}
        ws3.append([item_id, item.get("title", ""), item.get("status", "")])

    # --- Semantic: judged ambiguous rows ---
    ws4 = wb.create_sheet("Semantic")
    _write_header(ws4, ["source", "title", "target", "score", "verdict",
                        "reason"])
    for r in rows:
        if r["bucket"] == "ambiguous":
            sem = r.get("semantic") or {}
            ws4.append([r["source"], r["title"], r["target"], r["score"],
                        sem.get("verdict", ""), sem.get("reason", "")])

    return wb


def write_export(report: Dict, path: Path) -> Path:
    """Build and save the workbook to path; return the path."""
    wb = build_workbook(report)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
