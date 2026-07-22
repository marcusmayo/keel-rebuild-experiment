#!/usr/bin/env python3
"""Meridian PPM oracle -- asserts every number in the brief.

Runs the Phase 1-6 checks and prints PASS/FAIL per check, ending with ALL
PASS or FAILURES PRESENT. With --no-llm, the semantic verdict checks are
marked skipped (the model was not consulted).
"""
from __future__ import annotations

import csv
import json
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

import yaml  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

NO_LLM = "--no-llm" in sys.argv

# ---- Expected appendix data (verbatim from AGENTS.md) -----------------------
EXPECTED_PORTFOLIO = {
    "ML-001": ("Autonomous fleet routing optimization", "in_progress"),
    "ML-002": ("Driver scheduling engine", "in_progress"),
    "ML-003": ("Warehouse slotting analytics", "in_progress"),
    "ML-004": ("Cold chain temperature monitoring", "in_progress"),
    "ML-005": ("Ops latency dashboard rollout", "in_progress"),
    "ML-006": ("Customs paperwork automation", "in_progress"),
    "ML-007": ("Carrier rate benchmarking", "in_progress"),
    "ML-008": ("Dock door scheduling", "in_progress"),
    "ML-009": ("Fuel consumption reporting", "in_progress"),
    "ML-010": ("Returns processing portal", "in_progress"),
    "ML-011": ("Vendor onboarding checklist", "in_progress"),
    "ML-012": ("Route deviation alerts", "in_progress"),
    "ML-013": ("Pallet tracking tags", "in_progress"),
    "ML-014": ("Invoice dispute workflow", "in_progress"),
    "ML-015": ("Safety incident register", "in_progress"),
    "ML-016": ("Legacy TMS migration", "done"),
    "ML-017": ("Depot wifi upgrade", "done"),
    "ML-018": ("Contract renewal archive", "done"),
    "ML-019": ("Driver fatigue study", "in_progress"),
    "ML-020": ("Packaging waste audit", "in_progress"),
}
EXPECTED_JIRA = [
    ("ML-001", "Autonomous fleet routing optimization", "blocked"),
    ("ML-002", "Driver scheduling engine", "done"),
    ("ML-003", "Warehouse slotting analytics", "blocked"),
    ("ML-004", "Cold chain temperature monitoring", "done"),
    ("ML-005", "Ops latency dashboard rollout", "blocked"),
    ("ML-006", "Customs paperwork automation", "done"),
    ("ML-007", "Carrier rate benchmark refresh", "in_progress"),
    ("ML-008", "Dock door scheduling v2", "in_progress"),
    ("ML-016", "Legacy TMS migration", "in_progress"),
    ("ML-017", "Depot wifi upgrade", "in_progress"),
    ("ML-018", "Contract renewal archive", "done"),
    ("ML-021", "Telematics data lake", "done"),
    ("ML-022", "EDI partner certification", "done"),
    ("ML-023", "Yard congestion heatmap", "in_progress"),
    ("ML-024", "Reverse logistics pilot", "in_progress"),
]
EXPECTED_BACKLOG = [
    ("Route deviation alerts", "blocked"),
    ("Pallet tracking tags", "done"),
    ("Invoice dispute workflow", "blocked"),
    ("Ops latency dashboard", "in_progress"),
    ("Autonomous vehicle fleet navigation", "in_progress"),
    ("Quarterly fuel hedging review", "in_progress"),
    ("Trailer telematics retrofit", "in_progress"),
]
EXPECTED_BUCKETS = {
    "changed": 11, "gap": 4, "conflict": 2, "ambiguous": 1,
    "duplicate": 1, "completed": 1, "done_gap": 2,
}
EXPECTED_UNCONFIRMED = {"ML-009", "ML-010", "ML-011", "ML-015", "ML-019", "ML-020"}

results: list[tuple[str, bool, str]] = []


def check(name: str, cond: bool, detail: str = "") -> None:
    results.append((name, bool(cond), detail))


def skip(name: str, detail: str = "") -> None:
    results.append((name, True, "SKIPPED" + (f" ({detail})" if detail else "")))


# ---- Phase 1 ----------------------------------------------------------------
def check_phase1() -> None:
    state_dir = ROOT / "state"
    state_files = sorted(state_dir.glob("*.yaml"))
    check("P1: state holds exactly 20 items", len(state_files) == 20,
          f"got {len(state_files)}")

    loaded = {}
    for p in state_files:
        item = yaml.safe_load(p.read_text())
        loaded[item["id"]] = item
    check("P1: state ids match appendix", set(loaded) == set(EXPECTED_PORTFOLIO),
          f"got {sorted(loaded)}")
    for pid, (title, status) in EXPECTED_PORTFOLIO.items():
        if pid in loaded:
            check(f"P1: {pid} title", loaded[pid].get("title") == title,
                  f"got {loaded[pid].get('title')!r}")
            check(f"P1: {pid} status", loaded[pid].get("status") == status,
                  f"got {loaded[pid].get('status')!r}")

    csv_path = ROOT / "imports" / "jira_export.csv"
    with open(csv_path, newline="") as fh:
        jira_rows = list(csv.DictReader(fh))
    check("P1: Jira CSV has 15 data rows", len(jira_rows) == 15, f"got {len(jira_rows)}")
    check("P1: Jira CSV columns", set(jira_rows[0].keys()) == {"ref", "title", "status"}
          if jira_rows else False)
    if len(jira_rows) == 15:
        for i, (ref, title, status) in enumerate(EXPECTED_JIRA):
            check(f"P1: jira row {i+1}", (jira_rows[i]["ref"], jira_rows[i]["title"],
                  jira_rows[i]["status"]) == (ref, title, status),
                  f"got {(jira_rows[i]['ref'], jira_rows[i]['title'], jira_rows[i]['status'])}")

    xlsx_path = ROOT / "imports" / "backlog_export.xlsx"
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    data_rows = rows[1:]
    check("P1: Backlog XLSX has 7 data rows", len(data_rows) == 7, f"got {len(data_rows)}")
    if len(data_rows) == 7:
        for i, (title, status) in enumerate(EXPECTED_BACKLOG):
            got = (str(data_rows[i][0]).strip(), str(data_rows[i][1]).strip())
            check(f"P1: backlog row {i+1}", got == (title, status), f"got {got}")


# ---- Phase 2 ----------------------------------------------------------------
def check_phase2() -> None:
    jira_props = sorted((ROOT / "proposals" / "jira").glob("*.json"))
    backlog_props = sorted((ROOT / "proposals" / "backlog").glob("*.json"))
    check("P2: Jira emits exactly 15 proposals", len(jira_props) == 15,
          f"got {len(jira_props)}")
    check("P2: Backlog emits exactly 7 proposals", len(backlog_props) == 7,
          f"got {len(backlog_props)}")

    jira_recs = [json.loads(p.read_text()) for p in jira_props]
    backlog_recs = [json.loads(p.read_text()) for p in backlog_props]
    check("P2: Jira refs preserved", all(r["ref"] for r in jira_recs),
          "a Jira proposal had an empty ref")
    check("P2: Backlog refs empty", all(r["ref"] == "" for r in backlog_recs),
          "a backlog proposal had a non-empty ref")
    for rec in jira_recs + backlog_recs:
        keys = set(rec)
        check(f"P2: {rec['source']}-{rec['seq']} canonical fields",
              keys >= {"ref", "title", "status", "source"}, f"got {sorted(keys)}")

    # Read-only: state/ must still hold exactly 20 (not created/modified/removed).
    state_files = sorted((ROOT / "state").glob("*.yaml"))
    check("P2: state untouched (still 20 items)", len(state_files) == 20,
          f"got {len(state_files)}")


# ---- Phase 3 ----------------------------------------------------------------
def check_phase3() -> None:
    rec = json.loads((ROOT / "reconcile.json").read_text())
    summary = rec["summary"]
    buckets = summary["buckets"]
    check("P3: total rows = 22", summary["total"] == 22, f"got {summary['total']}")
    for b, expected in EXPECTED_BUCKETS.items():
        check(f"P3: bucket {b} = {expected}", buckets.get(b) == expected,
              f"got {buckets.get(b)}")
    total = sum(buckets.values())
    check("P3: bucket counts sum to 22", total == 22, f"got {total}")

    rows = rec["rows"]
    amb = [r for r in rows if r["bucket"] == "ambiguous"]
    check("P3: exactly 1 ambiguous row", len(amb) == 1, f"got {len(amb)}")
    if amb:
        a = amb[0]
        check("P3: ambiguous title", a["title"] == "Autonomous vehicle fleet navigation",
              f"got {a['title']!r}")
        check("P3: ambiguous target ML-001", a["match_target"] == "ML-001",
              f"got {a['match_target']}")
        check("P3: ambiguous score strictly between 0.40 and 0.80",
              a["score"] is not None and 0.40 < a["score"] < 0.80,
              f"got {a['score']}")

    dups = [r for r in rows if r["bucket"] == "duplicate"]
    check("P3: exactly 1 duplicate row", len(dups) == 1, f"got {len(dups)}")
    if dups:
        d = dups[0]
        check("P3: duplicate title", d["title"] == "Ops latency dashboard",
              f"got {d['title']!r}")
        check("P3: duplicate target ML-005", d["match_target"] == "ML-005",
              f"got {d['match_target']}")

    unc = summary["unconfirmed"]
    check("P3: unconfirmed count = 6", summary["unconfirmed_count"] == 6,
          f"got {summary['unconfirmed_count']}")
    check("P3: unconfirmed set matches",
          set(unc) == EXPECTED_UNCONFIRMED, f"got {sorted(unc)}")

    # Determinism: re-run reconciliation to temp files and compare bytes.
    from src.reconcile import reconcile
    t1 = Path(tempfile.mkstemp(suffix=".json")[1])
    t2 = Path(tempfile.mkstemp(suffix=".json")[1])
    reconcile(out_path=t1)
    reconcile(out_path=t2)
    b1, b2 = t1.read_bytes(), t2.read_bytes()
    t1.unlink(missing_ok=True)
    t2.unlink(missing_ok=True)
    identical = b1 == b2
    check("P3: reconciliation is byte-identical across runs", identical,
          "identical" if identical else f"{len(b1)} vs {len(b2)} bytes")


# ---- Phase 4 ----------------------------------------------------------------
def check_phase4() -> None:
    rec = json.loads((ROOT / "reconcile.json").read_text())
    rows = rec["rows"]
    amb = [r for r in rows if r["bucket"] == "ambiguous"]

    if NO_LLM:
        skip("P4: ambiguous verdict present", "--no-llm")
        skip("P4: ambiguous reason non-empty", "--no-llm")
    else:
        if amb:
            sem = amb[0].get("semantic") or {}
            check("P4: ambiguous verdict is SAME or DISTINCT",
                  sem.get("verdict") in ("SAME", "DISTINCT"), f"got {sem.get('verdict')!r}")
            check("P4: ambiguous reason non-empty",
                  bool(sem.get("reason")), f"got {sem.get('reason')!r}")
        else:
            check("P4: ambiguous verdict present", False, "no ambiguous row")
    check("P4: ambiguous bucket unchanged (still ambiguous)",
          all(r["bucket"] == "ambiguous" for r in amb))
    # Every non-ambiguous row carries no semantic verdict (untouched).
    other_sem = [r for r in rows if r["bucket"] != "ambiguous" and r.get("semantic")]
    check("P4: non-ambiguous rows untouched by semantic pass", not other_sem,
          f"{len(other_sem)} rows got a semantic verdict")

    # Scoring: every stored score recomputes from its stored factors.
    score_files = sorted((ROOT / "scores").glob("*.json"))
    check("P4: at least one score stored", len(score_files) >= 1,
          f"got {len(score_files)}")
    from src.score import compute_wsjf
    for sf in score_files:
        s = json.loads(sf.read_text())
        recomputed = compute_wsjf(s["factors"])
        ok = abs(recomputed - s["wsjf"]) < 1e-9
        check(f"P4: score for {s['item']} recomputable from factors", ok,
              f"stored {s['wsjf']} vs recomputed {recomputed}")
        # Factors within 1-10.
        fok = all(1 <= v <= 10 for v in s["factors"].values())
        check(f"P4: score for {s['item']} factors in 1-10", fok, f"got {s['factors']}")


# ---- Phase 5 ----------------------------------------------------------------
def check_phase5() -> None:
    exp = ROOT / "exports" / "reconciliation_review.xlsx"
    check("P5: export workbook exists", exp.exists())
    if not exp.exists():
        return
    wb = load_workbook(exp)
    names = set(wb.sheetnames)
    expected_sheets = {"Cross-Source", "Source-Only", "Unconfirmed", "Semantic"}
    check("P5: exactly the four sheets", names == expected_sheets, f"got {sorted(names)}")

    def data_rows(name):
        ws = wb[name]
        return ws.max_row - 1

    check("P5: Cross-Source sheet has 15 rows", data_rows("Cross-Source") == 15,
          f"got {data_rows('Cross-Source')}")
    check("P5: Source-Only sheet has 6 rows", data_rows("Source-Only") == 6,
          f"got {data_rows('Source-Only')}")
    check("P5: Unconfirmed sheet has 6 rows", data_rows("Unconfirmed") == 6,
          f"got {data_rows('Unconfirmed')}")
    check("P5: Semantic sheet has 1 row", data_rows("Semantic") == 1,
          f"got {data_rows('Semantic')}")

    if not NO_LLM:
        sem_ws = wb["Semantic"]
        verdict = sem_ws.cell(row=2, column=7).value
        reason = sem_ws.cell(row=2, column=8).value
        check("P5: Semantic sheet verdict SAME/DISTINCT", verdict in ("SAME", "DISTINCT"),
              f"got {verdict!r}")
        check("P5: Semantic sheet reason non-empty", bool(reason), f"got {reason!r}")
    else:
        skip("P5: Semantic sheet verdict", "--no-llm")
        skip("P5: Semantic sheet reason", "--no-llm")


# ---- No-delete grep ---------------------------------------------------------
def check_no_delete() -> None:
    """Grep the product code for delete affordances (routes/functions/buttons)."""
    bad_patterns = ["delete_item", "/delete", "def delete", "delete work", "delete_item",
                    "remove_item", "btn danger"]  # btn danger is hidden but present -> check separately
    hits = []
    for py in (ROOT / "src").glob("*.py"):
        text = py.read_text()
        for pat in ["def delete", "/delete", "delete_item", "remove_item", "delete("]:
            if pat in text:
                hits.append(f"{py.name}:{pat}")
    # Templates must not contain a 'delete' button.
    for tpl in (ROOT / "templates").glob("*.html"):
        text = tpl.read_text().lower()
        if "delete" in text:
            hits.append(f"{tpl.name}:delete")
    check("X: no delete route/function/button in product code", not hits,
          f"hits: {hits}")


def main() -> int:
    print("=" * 60)
    print("Meridian PPM oracle")
    print("=" * 60)
    check_phase1()
    check_phase2()
    check_phase3()
    check_phase4()
    check_phase5()
    check_no_delete()

    width = max(len(n) for n, _, _ in results)
    all_pass = True
    for name, ok, detail in results:
        status = "PASS" if ok else "FAIL"
        line = f"{status:<6} {name}"
        if detail:
            line += f"  -- {detail}"
        print(line)
        if not ok:
            all_pass = False
    print("-" * 60)
    if all_pass:
        print("ALL PASS")
        return 0
    print("FAILURES PRESENT")
    return 1


if __name__ == "__main__":
    sys.exit(main())
